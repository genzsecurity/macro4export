from openpyxl import load_workbook
import pandas as pd
import numpy as np
from openpyxl.utils.cell import get_column_letter
import os

def checkXSLX(path_to_file):

    # list of dicts declaration

    foundV = []

    # reading from excel file

    wb = load_workbook(filename = path_to_file)

    sheet_names = wb.sheetnames
    for sn in sheet_names:
        sheet_ranges = wb[sn]
        df1 = pd.DataFrame(sheet_ranges.values)
        df = df1.replace(np.nan, 'None', regex=True)

        # get rows number

        rowsNum=df.shape[0]

        # run through the sheet

        for r in range(0, rowsNum):
            for c in df.columns:

                # check if value is null

                cellV = df._get_value(r, c)
                if(cellV !='None'):

                    # get column name

                    cl=get_column_letter(c+1)

                    # append to list

                    foundV.append({'sheet':sheet_ranges, 'cell': cl, 'row': r+1, 'value':cellV})

    #print list and save as a file

    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    fullFilePath = os.path.join(desktop, "macro4export-output.txt")
    fileToWrite = open(fullFilePath, 'w')
    for f in foundV:
        print(f, file=fileToWrite)
        print(f)
    fileToWrite.close()


inp=input("Enter full file path:")
checkXSLX(inp)






